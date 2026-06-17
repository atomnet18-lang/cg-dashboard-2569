#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
convert_sharepoint_to_json.py
แปลงข้อมูลที่ผู้ปฏิบัติงานกรอกใน SharePoint List (export เป็น .xlsx/.csv)
ให้เป็น data.json สำหรับ Dashboard CG 2569 (GitHub Pages)

แนวคิดเดียวกับ Dashboard Compliance: จับคู่ด้วย AID (key เสถียร), ตัด field ภายในออก

โครงสร้าง List ที่คาดหวัง (1 แถว = 1 กิจกรรม):
  AID      : รหัสกิจกรรม เช่น ITA-01, NINE-38  (KEY — ห้ามแก้/ห้ามลบแถว)
  Plan     : ITA | คุณธรรม | Fraud | 9 หมวด    (อ้างอิงเฉยๆ)
  Section  : ชื่อหมวด (เฉพาะแผน 9 หมวด)
  ActID    : ลำดับกิจกรรมในแผน เช่น 1, 1.1
  Activity : ชื่อกิจกรรม
  Resp     : ผู้รับผิดชอบ
  Weight   : ค่าถ่วงน้ำหนัก
  ม.ค.-ธ.ค.: ผลจริงสะสม % รายเดือน 12 คอลัมน์
  Note     : หมายเหตุภายใน  ← ไม่ถูกเผยแพร่ (สคริปต์ตัดทิ้ง)

วิธีใช้:
  python3 convert_sharepoint_to_json.py <list_export.xlsx|csv> [data.json]

สคริปต์ใช้ data.json เดิมเป็น "แม่แบบโครงสร้าง" (ชื่อแผน/เป้าหมาย/หมวด/cum)
แล้วเติมเฉพาะผลจริงรายเดือนจาก List โดยจับคู่ด้วย AID
"""

import sys
import os
import json
import csv

MONTHS = ["ม.ค.","ก.พ.","มี.ค.","เม.ย.","พ.ค.","มิ.ย.","ก.ค.","ส.ค.","ก.ย.","ต.ค.","พ.ย.","ธ.ค."]
INTERNAL_FIELDS = {"note", "หมายเหตุ", "note (ใช้ภายใน)", "ปัญหา", "ปัญหา/อุปสรรค"}  # ตัดออกไม่เผยแพร่

HERE = os.path.dirname(os.path.abspath(__file__))


def norm(s):
    return str(s or "").strip()


def to_num(v):
    if v is None or v == "":
        return None
    try:
        return round(float(str(v).replace("%", "").strip()), 2)
    except (ValueError, TypeError):
        return None


def load_rows(path):
    """อ่าน export จาก SharePoint (.xlsx หรือ .csv) -> list of dict (header-keyed)"""
    ext = os.path.splitext(path)[1].lower()
    rows = []
    if ext in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        header = [norm(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not any(r):
                continue
            rows.append({header[i]: r[i] for i in range(len(header)) if i < len(r)})
    elif ext == ".csv":
        with open(path, encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                rows.append(row)
    else:
        raise SystemExit(f"รองรับเฉพาะ .xlsx หรือ .csv (ได้รับ {ext})")
    return rows


def month_index_map(header_keys):
    """หา mapping ชื่อคอลัมน์เดือน -> index 0..11 (รองรับชื่อเดือนตรงตัว)"""
    idx = {}
    for key in header_keys:
        k = norm(key)
        if k in MONTHS:
            idx[key] = MONTHS.index(k)
    return idx


def note_index_map(header_keys):
    """หา mapping คอลัมน์ 'สิ่งที่ทำ_<เดือน>' -> index 0..11"""
    idx = {}
    for key in header_keys:
        k = norm(key)
        if k.startswith("สิ่งที่ทำ_"):
            mon = k.split("_", 1)[-1]
            if mon in MONTHS:
                idx[key] = MONTHS.index(mon)
    return idx


def main():
    if len(sys.argv) < 2:
        print("ใช้งาน: python3 convert_sharepoint_to_json.py <list_export.xlsx|csv> [data.json]")
        sys.exit(1)
    src = sys.argv[1]
    dst = sys.argv[2] if len(sys.argv) > 2 else os.path.join(HERE, "data.json")

    # โหลดแม่แบบโครงสร้างจาก data.json เดิม
    tmpl_path = os.path.join(HERE, "data.json")
    if not os.path.exists(tmpl_path):
        raise SystemExit("ไม่พบ data.json (แม่แบบโครงสร้าง) ในโฟลเดอร์เดียวกัน")
    with open(tmpl_path, encoding="utf-8") as f:
        data = json.load(f)

    # index กิจกรรมตาม AID
    aid_index = {}
    for k, p in data["plans"].items():
        for a in p["acts"]:
            if a.get("aid"):
                a["a"] = [0]*12        # reset ผลจริง
                a["notes"] = [""]*12   # reset สิ่งที่ดำเนินการ
                aid_index[a["aid"]] = a

    rows = load_rows(src)
    if not rows:
        raise SystemExit("ไฟล์ export ว่างเปล่า")

    mcols = month_index_map(rows[0].keys())
    ncols = note_index_map(rows[0].keys())
    if len(mcols) < 12:
        missing = [m for m in MONTHS if m not in [norm(x) for x in rows[0].keys()]]
        print(f"คำเตือน: หาคอลัมน์เดือนไม่ครบ 12 (ขาด: {', '.join(missing)})")

    matched, unknown = 0, []
    for row in rows:
        aid = norm(row.get("AID") or row.get("aid"))
        if not aid:
            continue
        act = aid_index.get(aid)
        if not act:
            unknown.append(aid)
            continue
        for col, mi in mcols.items():
            v = to_num(row.get(col))
            if v is not None:
                act["a"][mi] = max(0, min(100, v))
        for col, mi in ncols.items():
            txt = norm(row.get(col))
            if txt:
                act["notes"][mi] = txt
        matched += 1

    # อัปเดตวันที่
    from datetime import date
    data["updated"] = date.today().strftime("%d/%m/2569")

    with open(dst, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"เขียน {dst} สำเร็จ: จับคู่ AID ได้ {matched} กิจกรรม")
    if unknown:
        print(f"คำเตือน: พบ AID ที่ไม่รู้จัก {len(unknown)} รายการ -> {', '.join(unknown[:10])}")
    print("หมายเหตุ: คอลัมน์ Note (หมายเหตุภายใน) ไม่ถูกนำออกเผยแพร่")


if __name__ == "__main__":
    main()
