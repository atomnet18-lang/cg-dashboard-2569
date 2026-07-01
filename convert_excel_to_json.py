#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
convert_excel_to_json.py
แปลงไฟล์แผน CG 4 แผน (.xlsx) เป็น data.json สำหรับ Dashboard บน GitHub Pages

วิธีใช้:
    python convert_excel_to_json.py "แผน_CG_4_แผน.xlsx" data.json

หลักการ:
- อ่านค่าเป้าหมายรายเดือน (12 เดือน) และค่าถ่วงน้ำหนักจากไฟล์แผนต้นฉบับ
- อ่าน "ผลการดำเนินงานจริง (ร้อยละ)" รายเดือนของแต่ละกิจกรรม (ถ้ายังไม่กรอก = 0)
- กรองเฉพาะ field ที่เปิดเผยสาธารณะได้ (ลำดับ/กิจกรรม/หน่วยงาน/น้ำหนัก/เป้าหมาย/ผลจริง)
  *ไม่* ดึงคอลัมน์ "ปัญหาหรืออุปสรรค" ออกไปเผยแพร่ เพื่อความปลอดภัยของข้อมูลภายใน
- เขียนผลเป็น data.json (UTF-8) ให้ index.html ดึงไปแสดง

โครงสร้างชีตที่รองรับ (ตามไฟล์แผน 2569):
    แถวหัวข้อแผนอยู่ด้านบน, ตารางกิจกรรมเริ่มหลังแถว header เดือน (ม.ค.-ธ.ค.)
    แต่ละกิจกรรมมี 2 แถว: แถว "เป้าหมาย (ร้อยละ)" และแถว "ผลการดำเนินงานจริง (ร้อยละ)"
"""

import sys
import json
import re
import openpyxl

# แมปชื่อชีต -> key + metadata สำหรับ dashboard
# รองรับทั้งไฟล์แผนต้นฉบับ (ชื่อยาว) และ template บันทึกผล (ชื่อสั้น)
SHEET_MAP = {
    # ไฟล์แผนต้นฉบับ
    "แผนประจำปี ITA 2569":     {"key": "ita",   "name": "ITA"},
    "แผนประจำปี คุณธรรม 2569":  {"key": "moral", "name": "คุณธรรม"},
    "แผนประจำปี Fraud 2569":   {"key": "fraud", "name": "Fraud"},
    "แผนประจำปี 9 หมวด 2569":   {"key": "nine",  "name": "9 หมวด"},
    # template บันทึกผล (ชื่อแท็บสั้น)
    "ITA":     {"key": "ita",   "name": "ITA"},
    "คุณธรรม":  {"key": "moral", "name": "คุณธรรม"},
    "Fraud":   {"key": "fraud", "name": "Fraud"},
    "9 หมวด":   {"key": "nine",  "name": "9 หมวด"},
}

MONTHS = ["ม.ค.","ก.พ.","มี.ค.","เม.ย.","พ.ค.","มิ.ย.","ก.ค.","ส.ค.","ก.ย.","ต.ค.","พ.ย.","ธ.ค."]


def clean(v):
    """แปลงค่าเซลล์เป็น string ที่สะอาด"""
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def num(v):
    """แปลงค่าเป็นตัวเลข; ถ้าว่าง/ไม่ใช่ตัวเลข คืน None"""
    if v is None or v == "":
        return None
    try:
        return round(float(v), 2)
    except (ValueError, TypeError):
        return None


def parse_meta(ws):
    """อ่าน metadata ของแผน (วัตถุประสงค์/เป้าหมาย/งบ/ตัวชี้วัด) จากแถวบนสุด"""
    meta = {"full": "", "budget": "-", "kpi": ""}
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        for cell in row:
            t = clean(cell)
            if not t:
                continue
            if meta["full"] == "" and "แผน" in t and "ประจำปี" in t:
                meta["full"] = t
            if t.startswith("งบประมาณ"):
                meta["budget"] = t.split(":", 1)[-1].strip() if ":" in t else "-"
            if t.startswith("ตัวชี้วัด"):
                meta["kpi"] = t.split(":", 1)[-1].strip() if ":" in t else t
    return meta


def find_month_cols(ws):
    """หาแถว header เดือน และคืน (row_index, [12 column indexes สำหรับ ม.ค.-ธ.ค.])"""
    for r in range(1, 16):
        vals = [clean(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        # แถวที่มี "ม.ค." และ "ธ.ค."
        if "ม.ค." in vals and "ธ.ค." in vals:
            cols = []
            for m in MONTHS:
                if m in vals:
                    cols.append(vals.index(m) + 1)  # 1-based
                else:
                    cols.append(None)
            return r, cols
    return None, None


def is_template_sheet(ws):
    """ตรวจว่าเป็นชีตแบบ template บันทึกผล (header เดือนอยู่แถว 4, มีคอลัมน์ผลจริงในเซลล์โดยตรง)"""
    for r in range(1, 8):
        vals = [clean(ws.cell(row=r, column=c).value) for c in range(1, min(ws.max_column, 20) + 1)]
        if "ลำดับ" in vals and "กิจกรรม" in vals and "น้ำหนัก" in vals:
            return r  # header row
    return None


def parse_template_sheet(ws, hdr_row):
    """อ่านชีต template: เป้าหมายดึงจาก comment ในเซลล์เดือน, ผลจริงดึงจากค่าในเซลล์"""
    meta = parse_meta(ws)
    # คอลัมน์: A ลำดับ, B กิจกรรม, C หน่วยงาน, D น้ำหนัก, E-P เดือน(12), Q ปัญหา, R-AC สิ่งที่ดำเนินการ(12)
    month_cols = list(range(5, 17))
    # หาคอลัมน์ "สิ่งที่ดำเนินการ <เดือน>" จากแถว header -> map เดือน:คอลัมน์
    note_cols = [None] * 12
    for c in range(1, ws.max_column + 1):
        h = clean(ws.cell(row=hdr_row, column=c).value)
        if h.startswith("สิ่งที่ดำเนินการ"):
            for i, m in enumerate(MONTHS):
                if m in h:
                    note_cols[i] = c
                    break
    acts = []
    cum = [0]*12
    r = hdr_row + 1
    while r <= ws.max_row:
        id_v = clean(ws.cell(row=r, column=1).value)
        name_v = clean(ws.cell(row=r, column=2).value)
        # แถวหมวด: merge เริ่มด้วย "หมวด"
        if name_v == "" and id_v.startswith("หมวด"):
            acts.append({"grp": id_v.replace("หมวด", "").strip()})
            r += 1; continue
        if id_v == "" and name_v == "":
            r += 1; continue
        if not (id_v or name_v):
            r += 1; continue
        # เป็นแถวกิจกรรม
        unit = clean(ws.cell(row=r, column=3).value) or "ฝบส."
        w = num(ws.cell(row=r, column=4).value) or 1
        targets, actuals = [], []
        for mc in month_cols:
            cell = ws.cell(row=r, column=mc)
            a_val = num(cell.value)
            actuals.append(a_val if a_val is not None else 0)
            # เป้าหมายจาก comment
            t_val = 0
            if cell.comment and cell.comment.text:
                m = re.search(r"(\d+(?:\.\d+)?)\s*%", cell.comment.text)
                if m:
                    t_val = float(m.group(1))
            targets.append(t_val)
        notes = [clean(ws.cell(row=r, column=nc).value) if nc else "" for nc in note_cols]
        acts.append({"id": id_v, "name": name_v, "unit": unit, "w": w,
                     "t": targets, "a": actuals, "notes": notes})
        r += 1
    # คำนวณเป้าหมายสะสมถ่วงน้ำหนักจาก targets ของกิจกรรม
    real = [a for a in acts if a.get("id")]
    tw = sum(a["w"] for a in real) or 1
    for i in range(12):
        cum[i] = round(sum(a["w"]*a["t"][i]/100 for a in real)/tw*100, 1)
    return {**meta, "cum": cum, "acts": acts}


def parse_plan(ws):
    meta = parse_meta(ws)
    hdr_row, month_cols = find_month_cols(ws)
    if hdr_row is None:
        return {**meta, "cum": [0]*12, "acts": []}

    acts = []
    cum_target = [0]*12

    # ค้นหาคอลัมน์ที่ระบุ "เป้าหมาย (ร้อยละ)" / "ผลการดำเนินงานจริง" (ปกติคอลัมน์ที่ 4)
    # และคอลัมน์ค่าถ่วงน้ำหนัก (หลังคอลัมน์ ธ.ค.)
    dec_col = month_cols[11] or month_cols[-1]
    weight_col = (dec_col + 1) if dec_col else None

    r = hdr_row + 1
    last_id = None
    last_name = ""
    last_unit = ""
    last_group = None

    while r <= ws.max_row:
        row_vals = [clean(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        joined = " ".join(row_vals)

        # แถวสรุปเป้าหมายสะสม
        if any(v.startswith("เป้าหมายสะสม") for v in row_vals):
            for i, mc in enumerate(month_cols):
                if mc:
                    n = num(ws.cell(row=r, column=mc).value)
                    cum_target[i] = n if n is not None else cum_target[i]
            r += 1
            continue
        if any("ผลการดำเนินงานจริงสะสม" in v for v in row_vals):
            r += 1
            continue

        col1 = row_vals[0]
        col2 = row_vals[1] if len(row_vals) > 1 else ""

        # ตรวจว่าเป็นแถวหัวข้อหมวด (สำหรับแผน 9 หมวด): มีลำดับเลขเดี่ยว + ชื่อหมวด ไม่มีค่าเป้าหมาย
        is_group = bool(re.fullmatch(r"\d+", col1)) and col2 and \
            not any(num(ws.cell(row=r, column=mc).value) is not None for mc in month_cols if mc)

        if is_group:
            last_group = col2
            acts.append({"grp": f"{col1} · {col2}"})
            r += 1
            continue

        # ตรวจแถว "เป้าหมาย (ร้อยละ)"
        marker = ""
        for c in range(1, 6):
            cv = clean(ws.cell(row=r, column=c).value)
            if "เป้าหมาย" in cv and "ร้อยละ" in cv:
                marker = "target"
            if "ผลการดำเนินงานจริง" in cv:
                marker = "actual"
        # บางครั้ง marker อยู่ในคอลัมน์ 4
        c4 = row_vals[3] if len(row_vals) > 3 else ""

        if "เป้าหมาย" in c4 and "ร้อยละ" in c4:
            # นี่คือแถวเป้าหมายของกิจกรรมหนึ่ง -> เก็บข้อมูลกิจกรรม
            act_id = col1 if col1 else (re.match(r"(\d+\.\d+)", col2).group(1) if re.match(r"(\d+\.\d+)", col2) else last_id)
            name = col2 if col2 else last_name
            unit = row_vals[2] if len(row_vals) > 2 else ""
            # ดึงเลขลำดับย่อยจากชื่อ (เช่น 1.1) ถ้าคอลัมน์ลำดับว่าง
            mid = re.match(r"\s*(\d+\.\d+)\s+(.*)", name)
            if mid:
                act_id = mid.group(1)
                name = mid.group(2)
            targets = []
            for mc in month_cols:
                n = num(ws.cell(row=r, column=mc).value) if mc else None
                targets.append(n if n is not None else 0)
            # ทำให้เป้าหมายเป็น cumulative-friendly (คงค่าล่าสุดเมื่อช่องถัดไปว่าง=0 หลังเริ่มแล้ว)
            seen = False
            for i in range(12):
                if targets[i] > 0:
                    seen = True
                elif seen and targets[i] == 0:
                    targets[i] = targets[i-1]
            weight = num(ws.cell(row=r, column=weight_col).value) if weight_col else None
            acts.append({
                "id": act_id or "",
                "name": clean(name),
                "unit": clean(unit) or "ฝบส.",
                "w": weight if weight is not None else 1,
                "t": targets,
                "a": [0]*12,   # ผลจริงตั้งต้น = รอบันทึก (ดึงจากแถวผลจริงถ้ามี)
            })
            last_id, last_name, last_unit = act_id, name, unit

        elif "ผลการดำเนินงานจริง" in c4 and acts and "id" in acts[-1]:
            # แถวผลจริง -> เติมค่าเข้ากิจกรรมล่าสุด
            for i, mc in enumerate(month_cols):
                n = num(ws.cell(row=r, column=mc).value) if mc else None
                if n is not None:
                    acts[-1]["a"][i] = n

        r += 1

    return {**meta, "cum": cum_target, "acts": acts}


def main():
    if len(sys.argv) < 3:
        print("ใช้งาน: python convert_excel_to_json.py <input.xlsx> <output.json>")
        sys.exit(1)
    src, dst = sys.argv[1], sys.argv[2]
    wb = openpyxl.load_workbook(src, data_only=True)

    plans = {}
    seen_keys = set()
    for sheet_name, info in SHEET_MAP.items():
        if sheet_name not in wb.sheetnames or info["key"] in seen_keys:
            continue
        ws = wb[sheet_name]
        hdr = is_template_sheet(ws)
        if hdr:
            parsed = parse_template_sheet(ws, hdr)
        else:
            parsed = parse_plan(ws)
        parsed["name"] = info["name"]
        plans[info["key"]] = parsed
        seen_keys.add(info["key"])

    # สร้าง aid ไม่ซ้ำต่อแผน (ใช้เป็น key เปิดหน้ารายละเอียดให้ถูกกิจกรรม)
    PREFIX = {"ita": "ITA", "moral": "MORAL", "fraud": "FRAUD", "nine": "NINE"}
    for key, p in plans.items():
        n = 0
        pfx = PREFIX.get(key, key.upper())
        for a in p["acts"]:
            if a.get("id"):
                n += 1
                a["aid"] = f"{pfx}-{n:02d}"

    # อ่านสรุปรายเดือนจากแท็บ "สรุปรายเดือน" (A=เดือน, B=ข้อความสรุป)
    reports = [""] * 12
    if "สรุปรายเดือน" in wb.sheetnames:
        wsr = wb["สรุปรายเดือน"]
        for r in range(1, wsr.max_row + 1):
            mname = clean(wsr.cell(row=r, column=1).value)
            if mname in MONTHS:
                val = wsr.cell(row=r, column=2).value
                reports[MONTHS.index(mname)] = str(val).strip() if val is not None else ""

    out = {
        "updated": "",   # เติมวันที่อัปเดตได้ภายหลัง
        "source": "ฝ่ายบริหารความเสี่ยงองค์กร (ฝบส.) การไฟฟ้านครหลวง",
        "plans": plans,
        "reports": reports,
    }
    with open(dst, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    n = sum(len([a for a in p["acts"] if a.get("id")]) for p in plans.values())
    print(f"เขียน {dst} สำเร็จ: {len(plans)} แผน, {n} กิจกรรม")


if __name__ == "__main__":
    main()
